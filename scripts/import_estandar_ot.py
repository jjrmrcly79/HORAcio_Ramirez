#!/usr/bin/env python3
# Carga el Excel "Estandar x Hora" (normalizado) + "OT PROCESO" (snapshot) a horacio.*
# Aplica primero sql/025_estandar_ot.sql, luego puebla partes / estandar_proceso / ordenes_trabajo.
# Idempotente (UPSERT por clave). Uso:
#   python3 scripts/import_estandar_ot.py "<estandar.xlsx>" "<ot.xlsx>" <YYYY-MM-DD snapshot> [--dry]
import os, sys, json, ssl, re, urllib.request, datetime
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
SK = ''
for ln in open(os.path.join(HERE, 'secrets.env')):
    if ln.startswith('SERVICE_ROLE_KEY='):
        SK = ln.split('=', 1)[1].strip()
ctx = ssl.create_default_context(); ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
DRY = '--dry' in sys.argv

def pg(q):
    req = urllib.request.Request("https://supabase.nexiasoluciones.com.mx/pg/query",
        data=json.dumps({"query": q}).encode(),
        headers={"apikey": SK, "Authorization": "Bearer " + SK, "Content-Type": "application/json"})
    return json.load(urllib.request.urlopen(req, context=ctx))

def esc(v):
    if v is None: return 'NULL'
    return "'" + str(v).replace("'", "''") + "'"

def num(v):
    if v is None or isinstance(v, str): return 'NULL'
    return repr(float(v))

def dt(v):
    if isinstance(v, (datetime.datetime, datetime.date)):
        return "'" + v.strftime('%Y-%m-%d') + "'"
    return 'NULL'

def norm_np(raw):
    if raw is None: return None
    s = str(raw).strip().upper()
    s = re.sub(r'_SMT$', '', s)
    return s.strip()

PROC_CODES = ['PP_481','PP_520','PP_411_481','PP_421','ENSAMBLE_MANUAL','WAVE_SOLDER',
              'SOLDEO_MANUAL','ICT','GRB','CONFORMAL','LIMPIEZA','FCT','ENSAMBLES',
              'PRUEBA_FCT','EMPAQUE']

def main():
    est_file, ot_file, snap = sys.argv[1], sys.argv[2], sys.argv[3]

    # --- DDL ---
    ddl = open(os.path.join(HERE, '..', 'sql', '025_estandar_ot.sql'), encoding='utf-8').read()
    if not DRY:
        pg(ddl); print("DDL 025 aplicado.")

    # =========================================================
    # ESTANDAR
    # =========================================================
    wb = openpyxl.load_workbook(est_file, data_only=True)
    ws = wb.worksheets[0]
    rows = list(ws.iter_rows(values_only=True))
    proc, hdr = rows[2], rows[3]
    std_cols = [i for i, h in enumerate(hdr) if h and 'Std' in str(h)]
    assert len(std_cols) == len(PROC_CODES), f"esperaba {len(PROC_CODES)} std cols, hay {len(std_cols)}"

    def block_range(idx):
        ci = std_cols[idx]
        start = next((j for j in range(ci, -1, -1) if proc[j]), ci)
        end = std_cols[idx + 1] if idx + 1 < len(std_cols) else len(hdr)
        # el bloque del proceso siguiente arranca en su group-label, no en su std col
        if idx + 1 < len(std_cols):
            nxt = std_cols[idx + 1]
            end = next((j for j in range(nxt, -1, -1) if proc[j]), nxt)
        return start, end

    blocks = [block_range(k) for k in range(len(std_cols))]
    data = [r for r in rows[4:] if len(r) > 4 and r[4] is not None]

    n_partes = n_std = 0
    for r in data:
        np_raw = r[4]
        npn = norm_np(np_raw)
        ens = str(r[5]).strip() if r[5] not in (None, '') else 'N/A'
        site, cliente, fam = r[1], r[2], r[3]
        if not DRY:
            res = pg(f"""INSERT INTO horacio.partes
                (numero_parte,no_parte_ensamble,numero_parte_raw,site,cliente,familia_modelo)
                VALUES({esc(npn)},{esc(ens)},{esc(np_raw)},{esc(site)},{esc(cliente)},{esc(fam)})
                ON CONFLICT (numero_parte,no_parte_ensamble) DO UPDATE SET
                  site=EXCLUDED.site, cliente=EXCLUDED.cliente, familia_modelo=EXCLUDED.familia_modelo,
                  numero_parte_raw=EXCLUDED.numero_parte_raw
                RETURNING id""")
            parte_id = res[0]['id'] if isinstance(res, list) else res['rows'][0]['id']
        else:
            parte_id = '(dry)'
        n_partes += 1

        for k, ci in enumerate(std_cols):
            v = r[ci] if len(r) > ci else None
            if v is None or isinstance(v, str) or v == 0:
                continue
            pzs = r[ci + 1] if len(r) > ci + 1 else None
            s, e = blocks[k]
            attrs = {}
            for j in range(s, e):
                if j in (ci, ci + 1): continue
                h = hdr[j] if j < len(hdr) else None
                val = r[j] if j < len(r) else None
                if h and val not in (None, ''):
                    key = re.sub(r'\s+', ' ', str(h).replace('\n', ' ')).strip()
                    attrs[key] = val if isinstance(val, (int, float)) else str(val)
            if not DRY:
                pg(f"""INSERT INTO horacio.estandar_proceso(parte_id,proceso,std_hr,pzs_turno,atributos)
                    VALUES({esc(parte_id)},{esc(PROC_CODES[k])},{num(v)},{num(pzs)},
                           {esc(json.dumps(attrs, ensure_ascii=False))}::jsonb)
                    ON CONFLICT (parte_id,proceso) DO UPDATE SET
                      std_hr=EXCLUDED.std_hr, pzs_turno=EXCLUDED.pzs_turno, atributos=EXCLUDED.atributos""")
            n_std += 1
    print(f"Estándar: {n_partes} partes · {n_std} filas estandar_proceso")

    # =========================================================
    # ORDENES DE TRABAJO
    # =========================================================
    wb2 = openpyxl.load_workbook(ot_file, data_only=True)
    ws2 = wb2.worksheets[0]
    otrows = [r for r in list(ws2.iter_rows(values_only=True))[1:] if r[0] is not None and r[1] is not None]
    n_ot = 0
    for r in otrows:
        ot = str(r[0]).strip()
        base, _, part = ot.rpartition('-')
        if not base: base, part = ot, None
        raw_np = r[1]
        es_smt = '_SMT' in str(raw_np).upper() or (part in ('02', '03'))
        npn = norm_np(raw_np)
        if not DRY:
            pg(f"""INSERT INTO horacio.ordenes_trabajo
                (orden_trabajo,orden_base,partida,es_smt,numero_parte,numero_parte_raw,descripcion,
                 tipo_ot,proceso_codigo,cant_ordenada,cant_terminada,fecha_orden,fecha_vence,
                 ventas,status_origen,fecha_snapshot)
                VALUES({esc(ot)},{esc(base)},{esc(part)},{str(es_smt).upper()},{esc(npn)},{esc(raw_np)},
                  {esc(r[2])},{esc(r[3])},{esc(r[4])},{num(r[5])},{num(r[6])},{dt(r[7])},{dt(r[8])},
                  {esc(r[9])},{esc(r[10])},{esc(snap)})
                ON CONFLICT (orden_trabajo,fecha_snapshot) DO UPDATE SET
                  cant_ordenada=EXCLUDED.cant_ordenada, cant_terminada=EXCLUDED.cant_terminada,
                  fecha_vence=EXCLUDED.fecha_vence, descripcion=EXCLUDED.descripcion""")
        n_ot += 1
    print(f"OT: {n_ot} órdenes (snapshot {snap}, estado_nexia=propuesta)")

if __name__ == '__main__':
    main()
