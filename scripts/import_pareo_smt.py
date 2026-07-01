#!/usr/bin/env python3
# ============================================================
# Horacio · loader del PAREO SMT↔final (semilla desde Excel de planeación)
# Siembra horacio.pareo_smt con fuente='excel_planeacion_2024-10'.
# La normalización REPLICA horacio.norm_np() (sql/040) para que empate con el export de OTs.
# Idempotente (ON CONFLICT). Los dos Excel viven en data/pareo/.
#
# Uso: python3 scripts/import_pareo_smt.py            # usa data/pareo/*.xlsx
#      python3 scripts/import_pareo_smt.py --dry      # no escribe, solo reporta
# ============================================================
import os, re, sys, json, ssl, urllib.request
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(HERE)
F_ENS = os.path.join(ROOT, 'data', 'pareo', 'ENSAMBLES.xlsx')
F_TAB = os.path.join(ROOT, 'data', 'pareo', 'Tabla de ensambles.xlsx')
FUENTE = 'excel_planeacion_2024-10'
DRY = '--dry' in sys.argv

def norm(p):
    """Réplica exacta de horacio.norm_np()."""
    s = ('' if p is None else str(p)).upper().strip()
    s = s.replace('(', '[').replace(')', ']')
    s = re.sub(r'_?SMT$', '', s)     # quita sufijo _SMT / SMT final
    s = re.sub(r'\s+', '', s)        # quita TODOS los espacios
    return s.rstrip('_')

def bad(x):
    return (not x) or x.upper() in ('N/A', 'NA')

def load_secrets():
    d = {}
    for ln in open(os.path.join(HERE, 'secrets.env')):
        ln = ln.strip()
        if ln and '=' in ln and not ln.startswith('#'):
            k, v = ln.split('=', 1); d[k] = v
    return d

def pg(query):
    sk = load_secrets()['SERVICE_ROLE_KEY']
    ctx = ssl.create_default_context(); ctx.check_hostname = False; ctx.verify_mode = ssl.CERT_NONE
    req = urllib.request.Request(
        'https://supabase.nexiasoluciones.com.mx/pg/query',
        data=json.dumps({'query': query}).encode(),
        headers={'apikey': sk, 'Authorization': 'Bearer ' + sk, 'Content-Type': 'application/json'})
    return json.loads(urllib.request.urlopen(req, context=ctx).read().decode())

# rows: dict (parte_smt, parte_final) -> {nivel, descripcion}
rows = {}
def add(smt, final, nivel, desc=''):
    ks, kf = norm(smt), norm(final)
    if bad(ks) or bad(kf) or ks == kf:
        return
    key = (ks, kf)
    if key not in rows:
        rows[key] = {'nivel': nivel, 'desc': (desc or '')[:120]}

# --- ENSAMBLES.xlsx: SMT(nombre) -> PT final (2 niveles) ---
ws = openpyxl.load_workbook(F_ENS, data_only=True).active
smt = None
for i, r in enumerate(ws.iter_rows(values_only=True)):
    if i == 0:
        continue
    c = [('' if x is None else str(x).strip()) for x in r]
    if c and c[0]:
        smt = c[0]
    pt = c[1] if len(c) > 1 else ''
    desc = c[2] if len(c) > 2 else ''
    if smt and pt:
        add(smt, pt, 'pt', desc)

# --- Tabla de ensambles: nivel SMT -> tarjeta PTH -> ensamble final (3 niveles) ---
ws2 = openpyxl.load_workbook(F_TAB, data_only=True)['240824']
lvl = None
for i, r in enumerate(ws2.iter_rows(values_only=True)):
    if i < 4:            # 0..2 vacías/título, 3 = encabezado
        continue
    c = [('' if x is None else str(x).replace('\n', ' ').strip()) for x in r]
    nivel_smt = c[2] if len(c) > 2 else ''
    ptht      = c[3] if len(c) > 3 else ''
    final     = c[7] if len(c) > 7 else ''
    if nivel_smt:
        lvl = nivel_smt
    if not bad(lvl):
        # Caso A: hay nivel SMT explícito → tarjeta PTH y final cuelgan de él
        if not bad(ptht):
            add(lvl, ptht, 'pth')
        if not bad(final):
            add(lvl, final, 'final')
    elif not bad(ptht) and not bad(final):
        # Caso B: nivel SMT en blanco → la "tarjeta" (col3) ES el subensamble (Led Board / Kit Wifi)
        add(ptht, final, 'final')

pairs = sorted(rows.items())
smts = sorted({k[0] for k in rows})
print(f'Parejas SMT↔final: {len(pairs)} · subensambles distintos: {len(smts)}')
for s in smts:
    n = sum(1 for k in rows if k[0] == s)
    print(f'  {s:<22} -> {n}')

if DRY:
    print('\n[dry] no se escribió nada.')
    sys.exit(0)

# upsert en lotes
def esc(s):
    return s.replace("'", "''")

vals = []
for (ks, kf), meta in pairs:
    vals.append(f"('{esc(ks)}','{esc(kf)}',{('NULL' if not meta['nivel'] else chr(39)+esc(meta['nivel'])+chr(39))},"
                f"{('NULL' if not meta['desc'] else chr(39)+esc(meta['desc'])+chr(39))},'{FUENTE}',true,'loader')")
sql = ("INSERT INTO horacio.pareo_smt (parte_smt,parte_final,nivel,descripcion,fuente,vigente,set_by_panel) VALUES "
       + ",".join(vals) +
       " ON CONFLICT (parte_smt,parte_final) DO UPDATE SET nivel=EXCLUDED.nivel, descripcion=EXCLUDED.descripcion, "
       "fuente=EXCLUDED.fuente, vigente=true, ts=now();")
pg(sql)
r = pg("SELECT count(*) n FROM horacio.pareo_smt WHERE vigente")
print(f"\n✓ Sembrado. Filas vigentes en pareo_smt: {r[0]['n']}")
